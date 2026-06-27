import os
import pandas as pd
import pdfplumber
import re
from openpyxl import load_workbook


class UniversalBillParser:
    def __init__(self, file_path):
        self.file_path = file_path
        self.file_ext = os.path.splitext(file_path)[1].lower()
        self.filename = os.path.basename(file_path)

    def parse(self):
        """核心分发器：自动嗅探文件格式并移交对应解析器"""
        print(f"🔍 正在嗅探文件: {self.filename}...")

        if self.file_ext == '.csv':
            return self._parse_csv()
        elif self.file_ext == '.pdf':
            return self._parse_pdf()
        elif self.file_ext in ['.xlsx', '.xls']:
            return self._parse_excel()
        else:
            raise ValueError(f"❌ 不支持的文件格式: {self.file_ext}")

    # ---------- CSV 解析（原有逻辑） ----------
    def _find_csv_header(self, encoding='utf-8'):
        """动态寻找真实表头，无视开头废话"""
        with open(self.file_path, 'r', encoding=encoding, errors='ignore') as f:
            for i, line in enumerate(f):
                if "交易时间" in line and ("金额" in line or "金额(元)" in line):
                    return i
        return 0

    def _parse_csv(self):
        """解析 CSV 账单 (兼容微信 & 支付宝)"""
        encodings_to_try = ['utf-8-sig', 'gbk', 'utf-8']
        df = None
        for enc in encodings_to_try:
            try:
                header_row_index = self._find_csv_header(encoding=enc)
                df = pd.read_csv(self.file_path, encoding=enc, skiprows=header_row_index)
                break
            except Exception:
                continue
        if df is None:
            raise Exception("❌ CSV 编码解析彻底失败，请检查文件！")

        df.columns = [col.strip() for col in df.columns]
        standard_data = []

        if "交易类型" in df.columns and "当前状态" in df.columns:
            print("🟢 识别为：微信 CSV 账单")
            df = df[~df['收/支'].astype(str).str.contains('中性交易|/', na=False)]
            for _, row in df.iterrows():
                try:
                    amt_str = str(row['金额(元)']).replace('¥', '').replace(',', '').strip()
                    amount = float(amt_str)
                    standard_data.append({
                        "date": str(row['交易时间']).strip(),
                        "amount": amount,
                        "type": "expense" if "支" in str(row['收/支']) else "income",
                        "raw_desc": f"对方:{row['交易对方']} | 商品:{row['商品']}"
                    })
                except Exception:
                    continue

        elif "交易分类" in df.columns and "交易状态" in df.columns:
            print("🔵 识别为：支付宝 CSV 账单")
            df = df[~df['收/支'].astype(str).str.contains('不计收支|/', na=False)]
            for _, row in df.iterrows():
                try:
                    amount = float(str(row['金额']).replace(',', '').strip())
                    standard_data.append({
                        "date": str(row['交易时间']).strip(),
                        "amount": amount,
                        "type": "expense" if "支" in str(row['收/支']) else "income",
                        "raw_desc": f"对方:{row['交易对方']} | 说明:{row['商品说明']}"
                    })
                except Exception:
                    continue

        return standard_data

    # ---------- PDF 解析（增强：支持系统导出纯文本） ----------
    def _parse_pdf(self):
        """解析 PDF：先尝试微信/支付宝表格，否则尝试纯文本格式"""
        standard_data = []
        try:
            with pdfplumber.open(self.file_path) as pdf:
                # 提取全文文本用于检测系统导出格式
                full_text = ""
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        full_text += page_text + "\n"

                # 检测是否系统导出的纯文本格式（每行以日期开头）
                lines = full_text.splitlines()
                system_format_count = 0
                for line in lines:
                    if re.match(r'^\d{4}-\d{2}-\d{2}\s+(支出|收入)\s+\S+\s+[\d.]+$', line.strip()):
                        system_format_count += 1
                        if system_format_count >= 2:
                            print("📄 识别为系统导出的纯文本账单 PDF")
                            return self._parse_system_export_text(full_text)

                # 否则按微信/支付宝表格处理
                first_page_text = pdf.pages[0].extract_text()
                is_wechat = "微信支付交易明细证明" in first_page_text
                is_alipay = "支付宝" in first_page_text

                for page in pdf.pages:
                    table = page.extract_table()
                    if not table:
                        continue
                    for row in table:
                        if not row or "交易时间" in str(row) or "金额" in str(row):
                            continue
                        row = [str(cell).replace('\n', '').strip() if cell else '' for cell in row]
                        try:
                            if is_wechat and len(row) >= 8:
                                if "中性交易" in row[3] or "其他" in row[3]:
                                    continue
                                amount = float(re.sub(r'[^\d.]', '', row[5]))
                                standard_data.append({
                                    "date": row[0][:19],
                                    "amount": amount,
                                    "type": "expense" if "支" in row[3] else "income",
                                    "raw_desc": f"对方:{row[7]} | 类型:{row[2]}"
                                })
                            elif is_alipay and len(row) >= 8:
                                if "不计收支" in row[0]:
                                    continue
                                amount = float(re.sub(r'[^\d.]', '', row[4]))
                                standard_data.append({
                                    "date": row[7][:19],
                                    "amount": amount,
                                    "type": "expense" if "支" in row[0] else "income",
                                    "raw_desc": f"对方:{row[1]} | 说明:{row[2]}"
                                })
                        except Exception:
                            continue
        except Exception as e:
            print(f"❌ PDF 解析失败: {e}")
        return standard_data

    def _parse_system_export_text(self, text: str) -> list:
        """
        解析系统导出的纯文本账单（每行：日期 类型 分类 金额）
        示例：2026-02-24 支出 K歌 80.00
        """
        lines = text.splitlines()
        bills = []
        pattern = re.compile(r'^(\d{4}-\d{2}-\d{2})\s+(支出|收入)\s+(\S+)\s+([\d.]+)$')
        for line in lines:
            line = line.strip()
            if not line:
                continue
            m = pattern.match(line)
            if m:
                date_str, type_cn, category, amount_str = m.groups()
                bill_type = 'income' if type_cn == '收入' else 'expense'
                amount_clean = re.sub(r'[^\d.]', '', amount_str)
                try:
                    amount = float(amount_clean)
                except ValueError:
                    continue
                bills.append({
                    'date': date_str,
                    'type': bill_type,
                    'category': category,
                    'amount': amount,
                    'raw_desc': f"{type_cn} {category} {amount_str}"
                })
        return bills

    # ---------- 新增 Excel 解析（支持 .xlsx / .xls） ----------
    def _parse_excel(self):
        """解析系统导出的 Excel 账单，自动识别列名（日期、类型、分类、金额、备注）"""
        wb = load_workbook(self.file_path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(cell).strip() if cell else "" for cell in rows[0]]
        data_rows = rows[1:]

        # 小写化表头便于匹配
        header_lower = [h.lower().strip() for h in header]

        # 定位各列索引
        date_idx = None
        type_idx = None
        category_idx = None
        amount_idx = None
        remark_idx = None

        for i, col in enumerate(header_lower):
            if '日期' in col or 'date' in col:
                date_idx = i
            elif '类型' in col or '收/支' in col or '收支' in col or 'type' in col:
                type_idx = i
            elif '分类' in col or 'category' in col or '类目' in col:
                category_idx = i
            elif '金额' in col or 'amount' in col or '交易金额' in col:
                amount_idx = i
            elif '备注' in col or '说明' in col or 'remark' in col or '描述' in col:
                remark_idx = i

        # 如果关键列缺失，使用启发式默认值
        if date_idx is None:
            # 默认第一列为日期
            date_idx = 0
        if amount_idx is None:
            # 尝试找带数字的列，或默认倒数第二列
            for i, col in enumerate(header_lower):
                if '金额' in col or 'amount' in col:
                    amount_idx = i
                    break
            if amount_idx is None:
                amount_idx = -1  # 最后一列
        if type_idx is None:
            # 尝试找“收/支”等列
            for i, col in enumerate(header_lower):
                if '收/支' in col or '类型' in col or '收支' in col:
                    type_idx = i
                    break
        if category_idx is None:
            for i, col in enumerate(header_lower):
                if '分类' in col or '类别' in col or '类目' in col:
                    category_idx = i
                    break
        if remark_idx is None:
            for i, col in enumerate(header_lower):
                if '备注' in col or '说明' in col or '描述' in col:
                    remark_idx = i
                    break

        bills = []
        for row in data_rows:
            if not row:
                continue

            # 日期
            date_val = row[date_idx] if date_idx < len(row) else None
            if date_val is None:
                continue
            try:
                date_str = pd.to_datetime(date_val).strftime("%Y-%m-%d")
            except:
                date_str = str(date_val).strip()

            # 金额
            amount_val = row[amount_idx] if amount_idx is not None and amount_idx < len(row) else None
            if amount_val is None:
                continue
            try:
                amount = float(amount_val)
            except:
                amt_str = str(amount_val).replace(',', '').replace('¥', '').strip()
                try:
                    amount = float(amt_str)
                except:
                    continue

            # 类型（收入/支出）
            type_val = row[type_idx] if type_idx is not None and type_idx < len(row) else None
            if type_val is None:
                # 根据金额正负推断
                if amount >= 0:
                    bill_type = "income"
                else:
                    bill_type = "expense"
                    amount = abs(amount)
            else:
                type_str = str(type_val).strip()
                if "支" in type_str or "支出" in type_str or "expense" in type_str.lower():
                    bill_type = "expense"
                    if amount > 0:
                        amount = amount
                elif "收" in type_str or "收入" in type_str or "income" in type_str.lower():
                    bill_type = "income"
                else:
                    if amount >= 0:
                        bill_type = "income"
                    else:
                        bill_type = "expense"
                        amount = abs(amount)

            # 分类
            category_val = row[category_idx] if category_idx is not None and category_idx < len(row) else None
            category = str(category_val).strip() if category_val else "其他"

            # 备注
            remark_val = row[remark_idx] if remark_idx is not None and remark_idx < len(row) else None
            remark = str(remark_val).strip() if remark_val else ""

            bills.append({
                "date": date_str,
                "amount": amount,
                "type": bill_type,
                "category": category,
                "raw_desc": remark
            })

        return bills

# ========== 测试入口 ==========
if __name__ == "__main__":
    # 测试文件路径，请按需修改
    test_file = r"../data/bill_test_data/账单导出_全量数据.pdf"
    parser = UniversalBillParser(test_file)
    result = parser.parse()
    print(f"✅ 解析出 {len(result)} 条记录")
    for item in result[:5]:
        print(item)